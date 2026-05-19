#!/usr/bin/env python3
"""Apply source-backed Secure Smart catalogue hierarchy/detail corrections safely.

Default is dry-run. Set SS_APPLY_CATALOG_SOURCE=1 to write files.

Rules:
- Preserve customer price fields and formula outputs exactly.
- Discomp brands: only replace category when a source breadcrumb/category is known.
- LifeSmart: use 2026-05-08 XLSX for title/category/description/stock-status text and
  explicit PCS/CTN from source XLSX only for carton quantity.
- Update all public data surfaces consistently: index, stage/enriched, product details.
"""
from __future__ import annotations

import json, os, re, hashlib
from pathlib import Path
from typing import Any

try:
    import openpyxl
except Exception as exc:  # pragma: no cover
    raise SystemExit(f"openpyxl is required: {exc}")

ROOT = Path(__file__).resolve().parents[1]
DATA = ROOT / 'assets' / 'catalog-data'
PRIVATE_DISCOMP = Path(os.environ.get('SS_DISCOMP_PRIVATE_SOURCE', '/srv/projects/secure-smart/data/secure_smart_catalog_enriched_discomp_stage_2026-05-02.json'))
DISCOMP_AUDIT = Path(os.environ.get('SS_DISCOMP_HIERARCHY_AUDIT', '/srv/projects/secure-smart/data/discomp_hierarchy_audit_2026-05-19.jsonl'))
DISCOMP_AUDITS = [Path(x) for x in os.environ.get('SS_DISCOMP_HIERARCHY_AUDITS', str(DISCOMP_AUDIT)).split(':') if x]
LIFESMART_FINAL = Path(os.environ.get('SS_LIFESMART_FINAL_XLSX', '/srv/projects/secure-smart/handoff/01-whatsapp-exports/extracted-kandiga-safe/0544-00006063-SecureSmart_All_Lists_WITH_LifeSmart_ZTE_USD_2026-05-08.xlsx'))
LIFESMART_CORE = Path(os.environ.get('SS_LIFESMART_CORE_XLSX', '/srv/projects/secure-smart/handoff/01-whatsapp-exports/extracted-kandiga-safe/0539-00005993-2026 Core smart home price with EAN final.xlsx'))
APPLY = os.environ.get('SS_APPLY_CATALOG_SOURCE') == '1'
REPORT = Path(os.environ.get('SS_CATALOG_SOURCE_REPORT', '/srv/projects/secure-smart/data/catalog_source_update_report_2026-05-19.json'))

PRICE_FIELDS = {'displayPriceUsd', 'fullCartonUnitPriceUsd', 'supplierPriceUsd', 'supplier_price_usd', 'secure_smart_price_raw', 'secure_smart_price_rounded', 'full_carton_unit_price_rounded'}
DISCOMP_BRANDS = {'Ubiquiti','MikroTik','Huawei','RF elements','Teltonika'}


def norm_sku(v: Any) -> str:
    return re.sub(r'\s+', ' ', str(v or '').strip()).upper()


def clean_text(s: Any) -> str:
    s = str(s or '').replace('\xa0',' ').replace('Â','')
    return re.sub(r'\s+', ' ', s).strip()


def is_bad_discomp_login_text(s: Any) -> bool:
    return 'requires your' in clean_text(s).lower() and 'log in' in clean_text(s).lower()


def load_json(path: Path):
    return json.loads(path.read_text(encoding='utf-8'))


def dump_json(path: Path, data: Any):
    # Keep compact JSON style used by catalogue data.
    path.write_text(json.dumps(data, ensure_ascii=False, separators=(',', ':')), encoding='utf-8')


def products_from(data: Any) -> list[dict]:
    if isinstance(data, list): return data
    if isinstance(data, dict) and isinstance(data.get('products'), list): return data['products']
    raise ValueError('Unsupported product JSON shape')


def load_discomp_private() -> dict[str, dict]:
    rows = products_from(load_json(PRIVATE_DISCOMP))
    return {norm_sku(p.get('sku')): p for p in rows if norm_sku(p.get('sku'))}


def load_discomp_audit_categories() -> dict[str, str]:
    out: dict[str, str] = {}
    for audit_path in DISCOMP_AUDITS:
        if not audit_path.exists():
            continue
        for line in audit_path.read_text(encoding='utf-8', errors='ignore').splitlines():
            try: r = json.loads(line)
            except Exception: continue
            if r.get('status') != 'ok': continue
            sku = norm_sku(r.get('sku'))
            labels = [clean_text(x.get('label')) for x in (r.get('breadcrumb') or []) if isinstance(x, dict) and clean_text(x.get('label'))]
            labels = [x for x in labels if x.lower() not in {'home', 'home page'}]
            if sku and labels:
                # Later audit files intentionally override older partial runs for the same SKU.
                out[sku] = ' / '.join(labels[:5])
    return out


def load_lifesmart_final() -> dict[str, dict]:
    wb = openpyxl.load_workbook(LIFESMART_FINAL, read_only=True, data_only=True)
    ws = wb['LIFESMART USD FINAL']
    headers = [str(c.value or '').strip() for c in ws[1]]
    rows = {}
    for vals in ws.iter_rows(min_row=2, values_only=True):
        d = {headers[i]: vals[i] for i in range(min(len(headers), len(vals)))}
        sku = norm_sku(d.get('secure_smart_sku'))
        if sku:
            rows[sku] = d
    return rows


def load_lifesmart_cartons() -> dict[str, int | None]:
    """Explicit PCS/CTN only. Missing => None; PCS/CTN=1 is kept as 1 but cartonAvailable false."""
    cartons: dict[str, int | None] = {}
    if not LIFESMART_CORE.exists(): return cartons
    wb = openpyxl.load_workbook(LIFESMART_CORE, read_only=True, data_only=True)
    ws = wb['smart home ']
    for row in range(10, ws.max_row + 1):
        sku = norm_sku(ws.cell(row, 3).value)
        notes = str(ws.cell(row, 13).value or '')
        if not sku: continue
        m = re.search(r'(\d+)\s*PCS\s*/\s*CTN', notes, re.I)
        cartons[sku] = int(m.group(1)) if m else None
    return cartons


def preferred_discomp_category(pub: dict, priv: dict | None, audited: dict[str, str]) -> str | None:
    sku = norm_sku(pub.get('sku'))
    if sku in audited:
        return audited[sku]
    if not priv: return None
    cat = clean_text(priv.get('category'))
    title = clean_text(priv.get('title'))
    # Avoid URL placeholder or one-off title-as-category dump; keep existing category if no real source category.
    if not cat or cat.startswith('http') or (title and cat == title):
        return None
    return cat


def build_lifesmart_description(src: dict) -> str:
    parts = [clean_text(src.get('description'))]
    attrs = []
    for label, key in [('Certificate','certificate'),('Adapter','adapter'),('Packaging','packaging'),('EAN','ean'),('HS code','hs_code'),('Package size','package_size'),('Weight','weight_kg'),('Notes','remark')]:
        val = clean_text(src.get(key))
        if val and val != '/':
            suffix = ' kg' if key == 'weight_kg' and not val.lower().endswith('kg') else ''
            attrs.append(f'{label}: {val}{suffix}')
    if attrs: parts.append(' | '.join(attrs))
    return '\n'.join([p for p in parts if p]).strip()


def candidate_updates(p: dict, *, private: dict[str,dict], audited: dict[str,str], lifesmart: dict[str,dict], cartons: dict[str,int|None], is_detail=False) -> dict[str, Any]:
    sku = norm_sku(p.get('sku'))
    brand = clean_text(p.get('brand'))
    updates: dict[str, Any] = {}

    if brand == 'LifeSmart' and sku in lifesmart:
        src = lifesmart[sku]
        title = clean_text(src.get('product_title'))
        cat = clean_text(src.get('category'))
        desc = build_lifesmart_description(src)
        stock = clean_text(src.get('stock_status'))
        if title: updates['title'] = title
        if cat: updates['category'] = cat
        if stock: updates['availability'] = stock
        # Explicit PCS/CTN only. Exact SKU first; fallback to model/base only if unique source match is explicit.
        ctn = cartons.get(sku)
        if ctn is None:
            base = norm_sku(src.get('model'))
            if base and base in cartons:
                ctn = cartons[base]
        updates['unitPerCarton'] = ctn
        updates['cartonAvailable'] = bool(ctn and ctn > 1)
        if is_detail and desc:
            updates['description'] = desc
        return updates

    if brand in DISCOMP_BRANDS:
        priv = private.get(sku)
        cat = preferred_discomp_category(p, priv, audited)
        if cat: updates['category'] = cat
        if priv:
            # Do not overwrite customer price fields. Use source title/description/warranty/carton only.
            title = clean_text((priv.get('discompEnrichment') or {}).get('title') or priv.get('title'))
            if is_bad_discomp_login_text(title):
                title = clean_text(priv.get('title'))
            if title and not is_bad_discomp_login_text(title): updates['title'] = title
            # Do not mass-normalize warranty in this hierarchy pass; it is outside the requested scope.
            # carton values from existing private stage/cartonMatch are source-derived; never infer from pack text.
            ctn = priv.get('unitPerCarton')
            if ctn is None and isinstance(priv.get('cartonMatch'), dict):
                ctn = priv['cartonMatch'].get('unitPerCarton')
            if ctn is not None:
                try: ctn = int(ctn)
                except Exception: ctn = None
                updates['unitPerCarton'] = ctn
                updates['cartonAvailable'] = bool(ctn and ctn > 1)
            if is_detail:
                desc = clean_text((priv.get('discompEnrichment') or {}).get('description') or priv.get('description'))
                if desc: updates['description'] = desc
        return updates

    return updates


def apply_updates(p: dict, updates: dict[str, Any]) -> dict[str, tuple[Any, Any]]:
    changes = {}
    before_prices = {k:p.get(k) for k in PRICE_FIELDS if k in p}
    for k,v in updates.items():
        if k in PRICE_FIELDS:
            raise AssertionError(f'Attempted price update {k}')
        if p.get(k) != v:
            changes[k] = (p.get(k), v)
            p[k] = v
    after_prices = {k:p.get(k) for k in PRICE_FIELDS if k in p}
    if before_prices != after_prices:
        raise AssertionError('Price fields changed')
    return changes


def main():
    private = load_discomp_private()
    audited = load_discomp_audit_categories()
    lifesmart = load_lifesmart_final()
    cartons = load_lifesmart_cartons()
    summary = {
        'apply': APPLY,
        'sources': {'privateDiscomp': str(PRIVATE_DISCOMP), 'discompAudits': [str(x) for x in DISCOMP_AUDITS], 'lifesmartFinal': str(LIFESMART_FINAL), 'lifesmartCore': str(LIFESMART_CORE)},
        'sourceCounts': {'privateDiscomp': len(private), 'discompAuditedCategories': len(audited), 'lifesmartRows': len(lifesmart), 'lifesmartCartonRows': sum(v is not None for v in cartons.values())},
        'files': {},
        'pricePreservationViolations': 0,
    }

    product_files = [DATA/'products.index.json', DATA/'products.stage.json', DATA/'products.enriched.stage.json']
    all_detail_changes = 0
    samples = []

    for path in product_files:
        if not path.exists(): continue
        data = load_json(path)
        rows = products_from(data)
        changed_rows = 0; changed_fields = {}
        for p in rows:
            updates = candidate_updates(p, private=private, audited=audited, lifesmart=lifesmart, cartons=cartons, is_detail=False)
            ch = apply_updates(p, updates)
            if ch:
                changed_rows += 1
                for k in ch: changed_fields[k] = changed_fields.get(k, 0) + 1
                if len(samples) < 20:
                    samples.append({'file': str(path.relative_to(ROOT)), 'sku': p.get('sku'), 'brand': p.get('brand'), 'changes': {k:[old,new] for k,(old,new) in ch.items()}})
        summary['files'][str(path.relative_to(ROOT))] = {'rows': len(rows), 'changedRows': changed_rows, 'changedFields': changed_fields}
        if APPLY and changed_rows:
            dump_json(path, data)

    detail_dir = DATA/'product-details'
    detail_files = list(detail_dir.glob('*.json')) if detail_dir.exists() else []
    detail_changed_fields = {}
    for path in detail_files:
        try: d = load_json(path)
        except Exception: continue
        target = d.get('product') if isinstance(d.get('product'), dict) else d
        if not isinstance(target, dict): continue
        updates = candidate_updates(target, private=private, audited=audited, lifesmart=lifesmart, cartons=cartons, is_detail=True)
        ch = apply_updates(target, updates)
        if ch:
            all_detail_changes += 1
            for k in ch: detail_changed_fields[k] = detail_changed_fields.get(k, 0) + 1
            if len(samples) < 20:
                samples.append({'file': str(path.relative_to(ROOT)), 'sku': target.get('sku'), 'brand': target.get('brand'), 'changes': {k:[old,new] for k,(old,new) in ch.items()}})
            if APPLY:
                dump_json(path, d)
    summary['files']['assets/catalog-data/product-details/*.json'] = {'rows': len(detail_files), 'changedRows': all_detail_changes, 'changedFields': detail_changed_fields}
    summary['samples'] = samples
    REPORT.parent.mkdir(parents=True, exist_ok=True)
    REPORT.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding='utf-8')
    print(json.dumps(summary, ensure_ascii=False, indent=2))
    print('report', REPORT)

if __name__ == '__main__':
    main()
